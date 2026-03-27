from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, send_file, Response, jsonify
from core.forensics.audit_ledger import AuditLedger
from pymongo import MongoClient
import os
import cv2
import torch
import uuid
import numpy as np
import threading
import time
import datetime
from werkzeug.utils import secure_filename
from datetime import datetime
from zoneinfo import ZoneInfo
from bson.objectid import ObjectId
import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from fpdf import FPDF

from torchreid.utils import FeatureExtractor
from ultralytics import YOLO

# --- Import Camera Module ---
import camera_module as cam
from camera_module import (
    gen_frames,
    start_cameras,
    frames_lock,
    frame_meta,
    latest_frames,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet

def generate_pdf(output_path, entries):
    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()

    elements = []

    for entry in entries:
        # Text
        text = f"Name: {entry['name']} | Camera: {entry['camera']}"
        elements.append(Paragraph(text, styles["Normal"]))
        elements.append(Spacer(1, 10))

        # Image
        if entry.get("image_path"):
            try:
                img = Image(entry["image_path"], width=200, height=150)
                elements.append(img)
                elements.append(Spacer(1, 20))
            except Exception as e:
                elements.append(Paragraph("Image load failed", styles["Normal"]))

    doc.build(elements)


def generate_excel(output_path, entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"

    # Headers
    ws.append(["Name", "Camera", "Image"])

    row = 2
    for entry in entries:
        ws.cell(row=row, column=1, value=entry["name"])
        ws.cell(row=row, column=2, value=entry["camera"])

        if entry.get("image_path"):
            try:
                img = XLImage(entry["image_path"])
                img.width = 100
                img.height = 75
                ws.add_image(img, f"C{row}")
            except Exception as e:
                ws.cell(row=row, column=3, value="Image error")

        row += 1

    wb.save(output_path)

app = Flask(__name__)
app.secret_key = "supersecretkey"
UPLOAD_FOLDER = "static/uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Initialize Audit Ledger
audit_ledger = AuditLedger("secure_audit.jsonl")

YOLO_WEIGHTS = "yolov11n.pt"
REID_MODEL_NAME = "osnet_x1_0"
REID_MODEL_PATH = os.path.expanduser("~/.cache/torch/checkpoints/osnet_x1_0_imagenet.pth")
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

print(f"[Main] Using device: {DEVICE}")

yolo_model = YOLO(YOLO_WEIGHTS)
yolo_model.to(DEVICE)

extractor = FeatureExtractor(
    model_name=REID_MODEL_NAME,
    model_path=REID_MODEL_PATH,
    device=DEVICE
)

CROP_FOLDER = "crops/unknown"

# --- Database Setup ---
client = MongoClient("mongodb://localhost:27017/")
db = client["person_reid"]
people_col = db["people"]
logs_col = db["logs"]
history_col = db["track_history"]
access_col = db["access_control"]
alerts_col = db["alerts"]

# ── Camera List — edit sources here ──────────────────────────────────────────
CAMERA_LIST = [
    (0, "Webcam"),
    (1, "Webcam2"),
    (2, "Webcam3"),
    # ("vid3.mp4",  "Lobby_Cam"),
    # ("vid4.mp4",  "Entrance_Cam"),
    # ("rtsp://user:pass@192.168.x.x/stream", "IP_Cam"),
]

# ── Start camera threads on app startup ──────────────────────────────────────
_camera_threads = []


# =============================================================================
# Helpers
# =============================================================================

def l2norm(v):
    return v / (np.linalg.norm(v) + 1e-12)

def extract_feature(img_path):
    img = cv2.imread(img_path)
    if img is None:
        return None
    results = yolo_model(img, verbose=False, classes=[0])
    best_crop = None
    if results and results[0].boxes:
        best_area = 0
        for r in results[0].boxes:
            x1, y1, x2, y2 = map(int, r.xyxy[0])
            w, h = x2 - x1, y2 - y1
            if w * h > best_area:
                best_area = w * h
                best_crop = img[y1:y2, x1:x2]
    final_img = best_crop if best_crop is not None else img
    if final_img.size == 0:
        return None
    img_rgb = cv2.cvtColor(final_img, cv2.COLOR_BGR2RGB)
    try:
        with torch.no_grad():
            feat_t = extractor([img_rgb])
        feat = feat_t[0].cpu().numpy().flatten()
        return l2norm(feat)
    except Exception as e:
        print(f"Error extracting feature: {e}")
        return None


# =============================================================================
# Existing Routes
# =============================================================================

@app.route("/")
def dashboard():
    return render_template("dashboard.html")


@app.route("/enrollment")
def enrollment():
    images = os.listdir(CROP_FOLDER) if os.path.exists(CROP_FOLDER) else []
    images = [f for f in images if f.lower().endswith((".jpg", ".png", ".jpeg"))]
    return render_template("enrollment.html", images=images)


@app.route("/enroll", methods=["POST"])
def enroll():
    name = request.form.get("name")
    role = request.form.get("role")
    selected_images = request.form.getlist("selected")
    uploaded_files = request.files.getlist("uploads")

    if not name or not role:
        return "Missing fields", 400

    features = []

    for img_file in selected_images:
        img_path = os.path.join(CROP_FOLDER, img_file)
        feat = extract_feature(img_path)
        if feat is not None:
            features.append(feat.tolist())

    for file in uploaded_files:
        if file and file.filename != "":
            filename = secure_filename(file.filename)
            save_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(save_path)
            feat = extract_feature(save_path)
            if feat is not None:
                features.append(feat.tolist())

    if features:
        now = datetime.now(ZoneInfo("Asia/Kolkata"))
        record = {
            "name": name,
            "role": role,
            "features": features,
            "registered_at": now.strftime('%Y-%m-%d %H:%M:%S')
        }
        people_col.insert_one(record)
        print(f"[+] Enrolled {name} ({role}) with {len(features)} images")
        audit_ledger.log("ENROLL_PERSON", {"name": name, "role": role})
        for img_file in selected_images:
            os.rename(os.path.join(CROP_FOLDER, img_file), f"crops/enrolled/{img_file}")

    return redirect(url_for("enrollment"))


@app.route("/people")
def people():
    all_people = list(people_col.find())
    return render_template("people.html", people=all_people)


@app.route("/edit/<person_id>", methods=["GET", "POST"])
def edit_person(person_id):
    person = people_col.find_one({"_id": ObjectId(person_id)})
    if not person:
        flash("Person not found.", "danger")
        return redirect(url_for("people"))

    if request.method == "POST":
        name = request.form.get("name")
        role = request.form.get("role")
        new_image_paths = person.get("images", [])
        uploaded_files = request.files.getlist("images")
        for file in uploaded_files:
            if file and file.filename != "":
                filename = secure_filename(f"{uuid.uuid4()}_{file.filename}")
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)
                new_image_paths.append(filepath)
                feat = extract_feature(filepath)
                if feat is not None:
                    people_col.update_one(
                        {"_id": ObjectId(person_id)},
                        {"$push": {"features": feat.tolist()}}
                    )
        people_col.update_one(
            {"_id": ObjectId(person_id)},
            {"$set": {"name": name, "role": role, "images": new_image_paths}}
        )
        audit_ledger.log("EDIT_PERSON", {"person_id": person_id, "new_name": name, "new_role": role})
        flash("Person updated successfully!", "success")
        return redirect(url_for("people"))

    if "images" not in person:
        person["images"] = []
    return render_template("edit_person.html", person=person)


@app.route("/delete/<person_id>")
def delete_person(person_id):
    person = people_col.find_one({"_id": ObjectId(person_id)})
    if person:
        for img_path in person.get("images", []):
            if os.path.exists(img_path):
                os.remove(img_path)
        people_col.delete_one({"_id": ObjectId(person_id)})
        audit_ledger.log("DELETE_PERSON", {"person_id": person_id, "name": person.get("name")})
        flash("Person deleted successfully!", "success")
    else:
        flash("Person not found.", "danger")
    return redirect(url_for("people"))


@app.route("/history", methods=["GET", "POST"])
def history():
    query_name = None
    results = []
    camera_filter = None
    start_date = None
    end_date = None
    sort_order = -1

    if request.method == "POST":
        from datetime import timedelta
        query_name = request.form.get("name", "").strip()
        camera_filter = request.form.get("camera", "").strip()
        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")
        sort_order = int(request.form.get("sort_order", "-1"))

        query = {}
        if query_name:
            query["person_name"] = {"$regex": f"^{query_name}$", "$options": "i"}
        if camera_filter:
            query["camera_name"] = {"$regex": f"^{camera_filter}$", "$options": "i"}
        if start_date or end_date:
            query["timestamp"] = {}
            if start_date:
                query["timestamp"]["$gte"] = datetime.strptime(start_date, "%Y-%m-%d")
            if end_date:
                query["timestamp"]["$lte"] = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

        results = list(history_col.find(query).sort("timestamp", sort_order))
        for r in results:
            r["formatted_time"] = r["timestamp"].strftime("%Y-%m-%d %H:%M:%S")
            if r.get("thumbnail") and not r["thumbnail"].startswith("/static/"):
                r["thumbnail"] = "/static/" + os.path.relpath(r["thumbnail"], "static").replace("\\", "/")

    all_cameras = sorted(set([x["camera_name"] for x in history_col.find({}, {"camera_name": 1})]))
    return render_template(
        "history.html",
        results=results,
        query_name=query_name,
        all_cameras=all_cameras,
        camera_filter=camera_filter,
        start_date=start_date,
        end_date=end_date,
        sort_order=sort_order,
    )


@app.route("/export/excel/<name>")
def export_excel(name):
    logs = list(history_col.find({"person_name": {"$regex": f"^{name}$", "$options": "i"}})
                            .sort("timestamp", -1))
    audit_ledger.log("DATA_EXPORT", {"type": "EXCEL", "subject": name})
    if not logs:
        return "No records found", 404

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "History"

    # Headers
    headers = ["Name", "Camera", "Timestamp", "Image"]
    ws.append(headers)

    row = 2
    for log in logs:
        ws.cell(row=row, column=1, value=log.get("person_name"))
        ws.cell(row=row, column=2, value=log.get("camera_name"))
        ws.cell(row=row, column=3, value=str(log.get("timestamp")))

        # Handle image
        thumb = log.get("thumbnail")
        if thumb:
            try:
                img_path = thumb.replace("/static/", "static/")
                if os.path.exists(img_path):
                    img = XLImage(img_path)
                    img.width = 100
                    img.height = 75
                    ws.add_image(img, f"D{row}")
                    ws.row_dimensions[row].height = 80
                else:
                    ws.cell(row=row, column=4, value="Missing")
            except:
                ws.cell(row=row, column=4, value="Error")

        row += 1

    wb.save(output)
    output.seek(0)

    filename = f"{name}_history_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export/pdf/<name>")
def export_pdf(name):
    logs = list(history_col.find({"person_name": {"$regex": f"^{name}$", "$options": "i"}})
                            .sort("timestamp", -1))
    if not logs:
        return "No records found", 404

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    elements = []

    title = f"Attendance History for {name}"
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 20))

    for log in logs:
        text = f"""
        Name: {log.get('person_name')}<br/>
        Camera: {log.get('camera_name')}<br/>
        Time: {log.get('timestamp')}<br/>
        Status: {log.get('status', 'Unknown')}
        """
        elements.append(Paragraph(text, styles["Normal"]))
        elements.append(Spacer(1, 10))

        # Handle image
        thumb = log.get("thumbnail")
        if thumb:
            try:
                img_path = thumb.replace("/static/", "static/")
                if os.path.exists(img_path):
                    img = Image(img_path, width=200, height=150)
                    elements.append(img)
                    elements.append(Spacer(1, 20))
                else:
                    elements.append(Paragraph("Image missing", styles["Normal"]))
            except:
                elements.append(Paragraph("Image error", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)

    filename = f"{name}_history_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")


@app.route("/logs/<path:filename>")
def logs(filename):
    return send_from_directory("thumbnails", filename)


@app.route("/access-control")
def access_control():
    access_docs = list(access_col.find())
    people = list(people_col.find({}, {"name": 1}))
    cameras = sorted(
        {d["camera_name"] for d in access_docs} |
        {"Camera_1", "Camera_2", "Camera_3", "Entrance_Cam", "Lobby_Cam", "Webcam", "cam4"} |
        {c for _, c in CAMERA_LIST}
    )
    return render_template("access_control.html", access_docs=access_docs, people=people, cameras=cameras)


@app.route("/access-control/update", methods=["POST"])
def update_access_control():
    cam_name = request.form.get("camera_name")
    allowed_people = request.form.getlist("allowed_people")
    if not cam_name:
        flash("Camera name missing!", "error")
        return redirect(url_for("access_control"))
    access_col.update_one(
        {"camera_name": cam_name},
        {"$set": {"allowed_people": allowed_people}},
        upsert=True
    )
    audit_ledger.log("UPDATE_ACCESS", {"camera": cam_name, "allowed": allowed_people})
    flash(f"Access list updated for {cam_name}", "success")
    return redirect(url_for("access_control"))


@app.route("/alerts", methods=["GET"])
def alerts():
    alerts_data = list(alerts_col.find().sort("timestamp", -1))
    return render_template("alerts.html", alerts=alerts_data)


@app.route("/export_alerts_excel")
def export_alerts_excel():
    alerts_data = list(alerts_col.find())
    if not alerts_data:
        return "No data available", 404
    df = pd.DataFrame(alerts_data)
    df.drop("_id", axis=1, inplace=True)
    file_path = "static/exports/alerts_log.xlsx"
    os.makedirs("static/exports", exist_ok=True)
    df.to_excel(file_path, index=False)
    return send_file(file_path, as_attachment=True)


@app.route("/export_alerts_pdf")
def export_alerts_pdf():
    alerts_data = list(alerts_col.find().sort("timestamp", -1))
    if not alerts_data:
        return "No data available", 404

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Alerts Log", ln=True, align="C")
    pdf.ln(10)

    for alert in alerts_data:
        pdf.multi_cell(
            0,
            10,
            txt=(
                f"Person: {alert.get('person_name')}\n"
                f"Camera: {alert.get('camera_name')}\n"
                f"Type: {alert.get('status')}\n"
                f"Timestamp: {alert.get('timestamp')}\n"
            ),
            border=1
        )
        pdf.ln(5)

    file_path = "static/exports/alerts_log.pdf"
    os.makedirs("static/exports", exist_ok=True)
    pdf.output(file_path)
    return send_file(file_path, as_attachment=True)


# =============================================================================
# /camera  —  Live Feed Routes
# =============================================================================

@app.route("/camera")
def camera():
    """Live camera dashboard page."""
    cam_names = [c for _, c in CAMERA_LIST]
    return render_template("camera.html", cameras=cam_names)


@app.route("/camera/stream/<cam_name>")
def camera_stream(cam_name: str):
    """MJPEG stream for a single camera tile."""
    valid_cams = [c for _, c in CAMERA_LIST]
    if cam_name not in valid_cams:
        return "Camera not found", 404
    return Response(
        gen_frames(cam_name),
        mimetype="multipart/x-mixed-replace; boundary=frame",
    )


@app.route("/camera/status")
def camera_status():
    """
    JSON: per-camera person count, identities, confidence, live alerts.
    Polled every second by the camera page.
    """
    with frames_lock:
        meta_snapshot = {k: dict(v) for k, v in frame_meta.items()}

    result = {}
    for _, cam_name in CAMERA_LIST:
        m = meta_snapshot.get(cam_name, {})
        result[cam_name] = {
            "count":      m.get("count", 0),
            "people":     m.get("people", []),
            "alerts":     m.get("alerts", []),
            "updated_at": m.get("updated_at", 0),
        }
    return jsonify(result)


@app.route("/camera/recent_logs")
def camera_recent_logs():
    """Last 20 detection log entries (used by camera sidebar)."""
    docs = list(
        history_col.find({}, {"_id": 0})
                   .sort("timestamp", -1)
                   .limit(20)
    )
    for d in docs:
        if isinstance(d.get("timestamp"), datetime):
            d["timestamp"] = d["timestamp"].strftime("%H:%M:%S")
    return jsonify(docs)


@app.route("/camera/recent_alerts")
def camera_recent_alerts():
    """Last 10 alerts (used by camera sidebar)."""
    docs = list(
        alerts_col.find({}, {"_id": 0})
                  .sort("timestamp", -1)
                  .limit(10)
    )
    for d in docs:
        if isinstance(d.get("timestamp"), datetime):
            d["timestamp"] = d["timestamp"].strftime("%H:%M:%S")
    return jsonify(docs)


# =============================================================================
# Entry Point
# =============================================================================

if __name__ == "__main__":
    os.makedirs(CROP_FOLDER, exist_ok=True)
    os.makedirs("crops/enrolled", exist_ok=True)

    _camera_threads = start_cameras(CAMERA_LIST)

    print(alerts_col.count_documents({}))
    print(alerts_col.find_one(sort=[("timestamp", -1)]))

    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True,
        threaded=True,
        use_reloader=False
    )